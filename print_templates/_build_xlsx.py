from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


OUT = Path(__file__).with_name("race_record_template.xlsx")

FONT_GOTHIC = "Hiragino Sans"
FONT_MINCHO = "Hiragino Mincho ProN"

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")


def set_range_border(ws, cell_range, outer=MEDIUM, inner=THIN):
    rows = list(ws[cell_range])
    min_row = rows[0][0].row
    max_row = rows[-1][0].row
    min_col = rows[0][0].column
    max_col = rows[0][-1].column

    for row in rows:
        for cell in row:
            left = outer if cell.column == min_col else inner
            right = outer if cell.column == max_col else inner
            top = outer if cell.row == min_row else inner
            bottom = outer if cell.row == max_row else inner
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def style_merged(ws, cell_range, value=None, font=None, alignment=None, fill=None):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    return cell


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "競漕記録"

    # Page setup: A4 landscape, fit on a single printed page.
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.39, right=0.39, top=0.39, bottom=0.39, header=0.12, footer=0.12
    )
    ws.print_options.horizontalCentered = True
    ws.print_area = "A1:M34"

    widths = {
        "A": 5.0,
        "B": 26.0,
        "C": 2.5,
        "D": 2.5,
        "E": 5.0,
        "F": 7.0,
        "G": 7.0,
        "H": 7.0,
        "I": 7.0,
        "J": 8.0,
        "K": 8.0,
        "L": 8.0,
        "M": 8.0,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in range(1, 34):
        ws.row_dimensions[row].height = 18

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24
    ws.row_dimensions[5].height = 24
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 25
    ws.row_dimensions[10].height = 27
    for row in range(11, 27, 2):
        ws.row_dimensions[row].height = 12
        ws.row_dimensions[row + 1].height = 24
    ws.row_dimensions[28].height = 27
    ws.row_dimensions[30].height = 28
    ws.row_dimensions[31].height = 28
    ws.row_dimensions[33].height = 22

    default_font = Font(name=FONT_GOTHIC, size=11)
    label_font = Font(name=FONT_GOTHIC, size=13, bold=True)
    header_font = Font(name=FONT_GOTHIC, size=14, bold=True)
    title_font = Font(name=FONT_GOTHIC, size=36, bold=True)
    large_font = Font(name=FONT_GOTHIC, size=16)
    fill_header = PatternFill("solid", fgColor="F2F2F2")

    center = Alignment(horizontal="center", vertical="center")
    left_center = Alignment(horizontal="left", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    shrink_center = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
    shrink_left_center = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)

    for row in ws.iter_rows(min_row=1, max_row=33, min_col=1, max_col=13):
        for cell in row:
            cell.font = default_font
            cell.alignment = center

    style_merged(
        ws,
        "A1:D1",
        "第17回全日本マスターズレガッタ",
        Font(name=FONT_GOTHIC, size=14, bold=True),
        left_center,
    )
    style_merged(ws, "E1:H2", "競漕記録", title_font, center)
    style_merged(ws, "J1:M1", "print", Font(name=FONT_GOTHIC, size=10), left_center)

    ws["J2"] = "競漕委員長"
    ws["K2"] = "審判長"
    style_merged(ws, "L2:M2", "主席判定員", label_font, wrap_center, fill_header)
    for cell in ws["J2:M2"][0]:
        cell.font = label_font
        cell.alignment = wrap_center
        cell.fill = fill_header
    style_merged(ws, "J3:J5")
    style_merged(ws, "K3:K5")
    style_merged(ws, "L3:M5")
    set_range_border(ws, "J2:M5")

    ws["A5"] = None

    style_merged(ws, "A7:B7", "Race No.", label_font, center)
    style_merged(ws, "C7:D7", "レース時間", label_font, center)
    style_merged(ws, "E7:J7", "種目名", label_font, left_center)
    style_merged(ws, "K7:M7", "", label_font, left_center)
    style_merged(ws, "A8:B8", "134", label_font, center)
    style_merged(ws, "C8:D8", "2026/05/23　07:00", label_font, shrink_center)
    style_merged(ws, "E8:J8", "", label_font, shrink_left_center)
    style_merged(ws, "K8:M8", "", label_font, shrink_left_center)
    set_range_border(ws, "A7:M8")

    headers = [("A10", "順位"), ("B10:D10", "クルー名"), ("E10", "レーン"), ("F10:G10", "500m"), ("H10:I10", "1000m"), ("J10:M10", "備考")]
    for ref, text in headers:
        if ":" in ref:
            style_merged(ws, ref, text, header_font, center, fill_header)
        else:
            ws[ref] = text
            ws[ref].font = header_font
            ws[ref].alignment = center
            ws[ref].fill = fill_header

    for lane in range(8):
        top = 11 + lane * 2
        bottom = top + 1
        style_merged(ws, f"A{top}:A{bottom}", "", large_font, center)
        style_merged(ws, f"B{top}:D{bottom}", "", large_font, shrink_center)
        style_merged(ws, f"E{top}:E{bottom}", "", large_font, center)
        style_merged(ws, f"F{top}:G{bottom}", "", large_font, center)
        style_merged(ws, f"H{top}:I{bottom}", "", large_font, center)
        style_merged(ws, f"J{top}:M{bottom}", "", default_font, center)

    set_range_border(ws, "A10:M26")

    style_merged(ws, "A28:B28", "天候:", label_font, left_center)
    style_merged(ws, "C28:D28", "", default_font, center)
    style_merged(ws, "E28:F28", "風向:", label_font, left_center)
    style_merged(ws, "G28:H28", "", default_font, center)
    style_merged(ws, "I28:J28", "風速:", label_font, left_center)
    style_merged(ws, "K28:M28", "", default_font, center)
    set_range_border(ws, "A28:M28")

    style_merged(ws, "A30:B31", "備考:", label_font, left_center)
    style_merged(ws, "C30:M31", "", default_font, center)
    set_range_border(ws, "A30:M31")

    style_merged(
        ws,
        "J33:M33",
        "石川県ボート協会",
        Font(name=FONT_GOTHIC, size=13, bold=True),
        Alignment(horizontal="right", vertical="center"),
    )

    ws.freeze_panes = None
    wb.save(OUT)


if __name__ == "__main__":
    main()
