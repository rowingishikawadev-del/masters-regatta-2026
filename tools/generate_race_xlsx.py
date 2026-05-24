#!/usr/bin/env python3
import argparse
import json
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[1]
MASTER_JSON = ROOT / "data" / "master.json"
TEMPLATE_XLSX = ROOT / "print_templates" / "race_record_template.xlsx"
SHRINK_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=True)
SHRINK_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=True)
ROUND_LABELS = {
    "FA": "決勝A",
}


DUMMY_RESULTS = {
    1: [
        {"rank": 1, "crew_name": "ＲＣ神戸　なでしこＯｈバーン", "lane": 4, "time_500": "1:38.20", "time_1000": "3:24.50", "note": ""},
        {"rank": 2, "crew_name": "愛知東郷ボートクラブ", "lane": 2, "time_500": "1:42.50", "time_1000": "3:31.20", "note": ""},
        {"rank": 3, "crew_name": "瀬田漕艇クラブ", "lane": 5, "time_500": "1:44.00", "time_1000": "3:34.80", "note": ""},
        {"rank": 4, "crew_name": "Ｅ．Ｒ．Ｃ．Ｃ　Ｗ４Ｘ＋　Ｄ", "lane": 1, "time_500": "1:45.30", "time_1000": "3:38.10", "note": ""},
        {"rank": 5, "crew_name": "宮ヶ瀬", "lane": 3, "time_500": "1:47.10", "time_1000": "3:42.60", "note": ""},
        {"rank": 6, "crew_name": "浜寺マスターズアマゾネスＢ", "lane": 6, "time_500": "1:51.80", "time_1000": "3:48.30", "note": ""},
    ]
}


def find_race(master, race_no):
    for race in master["schedule"]:
        if int(race["race_no"]) == race_no:
            return race
    raise ValueError(f"race_no {race_no} not found in {MASTER_JSON}")


def format_race_datetime(race):
    year, month, day = [int(part) for part in race["date"].split("/")]
    hour, minute = [int(part) for part in race["time"].split(":")]
    return f"{year:04d}/{month:02d}/{day:02d}　{hour:02d}:{minute:02d}"


def copy_top_left_style(ws, source, targets):
    style_source = ws[source]
    for target in targets:
        cell = ws[target]
        cell.font = copy(style_source.font)
        cell.alignment = copy(style_source.alignment)
        cell.fill = copy(style_source.fill)
        cell.border = copy(style_source.border)


def write_results(ws, results):
    row_pairs = [(11 + idx * 2, 12 + idx * 2) for idx in range(8)]
    for idx, (top, _bottom) in enumerate(row_pairs):
        result = results[idx] if idx < len(results) else {}
        ws[f"A{top}"] = result.get("rank", "")
        ws[f"B{top}"] = result.get("crew_name", "")
        ws[f"B{top}"].alignment = SHRINK_CENTER
        ws[f"E{top}"] = result.get("lane", "")
        ws[f"F{top}"] = result.get("time_500", "")
        ws[f"H{top}"] = result.get("time_1000", "")
        ws[f"J{top}"] = result.get("note", "")


def generate(race_no):
    with MASTER_JSON.open(encoding="utf-8") as f:
        master = json.load(f)

    race = find_race(master, race_no)
    round_label = ROUND_LABELS.get(race["round"], race["round"])

    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active
    ws.print_area = "A1:M34"

    ws["A1"] = master["tournament"]["race_name"]
    ws["E1"] = "競漕記録"
    ws["J1"] = f"print {datetime.now().strftime('%Y/%m/%d %H:%M:%S')}"
    ws["A8"] = race_no
    ws["C8"] = format_race_datetime(race)
    ws["E8"] = race["event_name"]
    ws["K8"] = round_label
    ws["J33"] = "石川県ボート協会"

    copy_top_left_style(ws, "A8", ["C8", "E8", "K8"])
    ws["C8"].alignment = SHRINK_CENTER
    ws["E8"].alignment = SHRINK_LEFT_CENTER
    ws["K8"].alignment = SHRINK_LEFT_CENTER
    write_results(ws, DUMMY_RESULTS.get(race_no, []))

    output = ROOT / "print_templates" / f"{race_no}.xlsx"
    wb.save(output)
    return output


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("race_no", type=int)
    args = parser.parse_args()
    print(generate(args.race_no))


if __name__ == "__main__":
    main()
